from openpyxl import load_workbook
import re


def replace(filename : str):
    try:
        wb = load_workbook(filename)
        for sheet in wb.sheetnames:
            table = wb[sheet]
            for row in table.columns:
                for cell in row:
                    val = cell.value if cell.value else ''
                    val = str(val)
                    name = re.search('Name.+', val)
                    address = re.search('Address.+', val)
                    if name:
                        cell.value = cell.value.replace(name.group(), '')
                    if address:
                        cell.value = cell.value.replace(address.group(), '')
                    
        
        wb.save(f'./media/{filename}')
        print(f'{filename} successfully saved in results/')
    except Exception as e:
        print(f'Error {e} occured')
